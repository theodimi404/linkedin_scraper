import getpass
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from tqdm import tqdm

ascii_art = f''' 
_  _         _               _  _
| |(_) _ __  | | __  ___   __| |(_) _ __    ___   ___  _ __   __ _  _ __    ___  _ __
| || || '_ \ | |/ / / _ \ / _` || || '_ \  / __| / __|| '__| / _` || '_ \  / _ \| '__|
| || || | | ||   < |  __/| (_| || || | | | \__ \| (__ | |   | (_| || |_) ||  __/| |
|_||_||_| |_||_|\_\ \___| \__,_||_||_| |_| |___/ \___||_|    \__,_|| .__/  \___||_|
                                                                   |_|
'''

print(ascii_art)

# specifies the path to the chromedriver.exe
driver = webdriver.Chrome()

# driver.get method() will navigate to a page given by the URL address
driver.get('https://www.linkedin.com')

# locate username form by_class_name
username = driver.find_element_by_id('session_key')

# send_keys() to simulate key strokes
input_username = input('USERNAME: ')
username.send_keys(input_username)

time.sleep(0.5)

# locate password form by_class_name
password = driver.find_element_by_id('session_password')

# send_keys() to simulate key strokes
input_password = getpass.getpass('PASSWORD: ')
password.send_keys(input_password)
time.sleep(0.5)

log_in_button = driver.find_element_by_class_name('sign-in-form__submit-button')
log_in_button.click()
time.sleep(10)

# Finding Network button and clicking on it
network_button = driver.find_element_by_link_text('My Network')
network_button.click()
time.sleep(10)

# Finding Connections button and click on it
connections_button = WebDriverWait(driver, 50).until(
    ec.presence_of_element_located((By.CLASS_NAME, "mn-community-summary__entity-info"))
)
connections_button.click()
time.sleep(5)
# TODO click the END key as many times as needed
driver.find_element_by_xpath('//body').send_keys(Keys.CONTROL + Keys.END)
time.sleep(2)

# Saving all connections in a list and extracting all profile links from it
page = driver.page_source
soup = BeautifulSoup(page, "html.parser")

connections = soup.find_all('li', {'class': 'mn-connection-card artdeco-list ember-view'})

links = []
for connection in connections:
    link = [a['href'] for a in connection.find_all('a', href=True)]
    links.append(link[0])


url_profiles = []

for link in links:
    url = "https://www.linkedin.com" + link
    url_profiles.append(url)

# Initializing the .xlsx file  
book = Workbook()
sheet = book.active
row = 1

for url in tqdm(url_profiles[4:]):
    time.sleep(5)

    driver.get(url)
    time.sleep(1)
    driver.find_element_by_xpath('//body').send_keys(Keys.SPACE)
    time.sleep(10)
    profile = driver.page_source

    time.sleep(3)

    profile_info = BeautifulSoup(profile, "html.parser")
    # print(profile_info)

    # Check if the connection is currently working
    experience = profile_info.find('div', {'id': 'oc-background-section'})
    for i in range(3):
        if not experience:
            driver.find_element_by_xpath('//body').send_keys(Keys.SPACE)
            time.sleep(10)
            profile = driver.page_source
            time.sleep(3)
            profile_info = BeautifulSoup(profile, "html.parser")
            experience = profile_info.find('div', {'id': 'oc-background-section'})
    try:
        # print(experience)
        last_experience = experience.find('li', {
            'class': 'pv-entity__position-group-pager pv-profile-section__list-item ember-view'})
        last_experience_dates = last_experience.find(
            'h4', {'class': 'pv-entity__date-range t-14 t-black--light t-normal'}
        )
        dates = last_experience_dates.find_all('span')
        dates = dates[1]
        if 'Present' in dates.text:
            unemployed = 'False'
        else:
            unemployed = 'True'
    except:
        unemployed = 'didnt found'

    time.sleep(1)

    contact_info = profile_info.find('a', {'data-control-name': 'contact_see_more'})

    contact_info_click = "https://www.linkedin.com" + contact_info['href']

    driver.get(contact_info_click)
    profile_info = driver.page_source

    soup1 = BeautifulSoup(profile_info, "html.parser")

    profile = soup1.find('section', {'class': 'pv-contact-info__contact-type ci-vanity-url'})

    email = soup1.find('section', {'class': 'pv-contact-info__contact-type ci-email'})

    connected = soup1.find('section', {'class': 'pv-contact-info__contact-type ci-connected'})

    birthday_info = soup1.find('section', {'class': 'pv-contact-info__contact-type ci-birthday'})

    if profile:
        profile_header = profile.find('header', {'class': 'pv-contact-info__header t-16 t-black t-bold'})
        profile_span = profile.find('a', {'class': 'pv-contact-info__contact-link link-without-visited-state t-14'})

        profile_info = [profile_header.text, profile_span['href']]

        sheet.cell(row=row, column=1).value = profile_span['href']
        # print(profile_info)
    if email:
        email_header = email.find('header', {'class': 'pv-contact-info__header t-16 t-black t-bold'})
        email_span = email.find('a', {'class': 'pv-contact-info__contact-link link-without-visited-state t-14'})

        email = [email_header.text, email_span.text]

        sheet.cell(row=row, column=2).value = email_span.text
        # print(email)
    if connected:
        connected_header = connected.find('header', {'class': 'pv-contact-info__header t-16 t-black t-bold'})
        connected_span = connected.find('span', {'class': 'pv-contact-info__contact-item t-14 t-black t-normal'})

        connected_info = [connected_header.text, connected_span.text]

        sheet.cell(row=row, column=3).value = connected_span.text
        # print(connected_info)
    if birthday_info:
        birthday_info_header = birthday_info.find('header', {'class': 'pv-contact-info__header t-16 t-black t-bold'})
        birthday_info_span = birthday_info.find('span',
                                                {'class': 'pv-contact-info__contact-item t-14 t-black t-normal'})

        birthdays = [birthday_info_header.text, birthday_info_span.text]

        sheet.cell(row=row, column=4).value = birthday_info_span.text
        # print(birthdays)
    if unemployed == 'True':
        # print(unemployed)
        sheet.cell(row=row, column=5).value = 'unemployed'
    elif unemployed == 'False':
        sheet.cell(row=row, column=5).value = 'employed'
    else:
        sheet.cell(row=row, column=5).value = 'didnt found'

    row += 1
    
driver.quit()
book.save("data_from_connections.xlsx")
